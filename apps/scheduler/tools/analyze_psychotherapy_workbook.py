from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


INPUT = Path(r"C:\Users\Rodrigo\Downloads\Controle_Sessoes_Psicoterapia.xlsx")


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def compact_row(row):
    return [cell_to_text(v) for v in row]


def infer_header_row(ws, max_scan=25):
    best = (0, 1)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        texts = [cell_to_text(v) for v in values]
        nonempty = sum(1 for t in texts if t)
        words = sum(1 for t in texts if re.search(r"[A-Za-zÀ-ÿ]", t))
        if nonempty + words > best[0]:
            best = (nonempty + words, r)
    return best[1]


def normalize_key(text):
    return re.sub(r"\s+", " ", text.strip().lower())


def analyze_sheet(ws_values, ws_formulas):
    header_row = infer_header_row(ws_formulas)
    headers = [cell_to_text(ws_formulas.cell(header_row, c).value) or f"Coluna {c}" for c in range(1, ws_formulas.max_column + 1)]

    rows = []
    formula_rows = []
    for r in range(header_row + 1, ws_formulas.max_row + 1):
        formula_vals = [ws_formulas.cell(r, c).value for c in range(1, ws_formulas.max_column + 1)]
        cached_vals = [ws_values.cell(r, c).value for c in range(1, ws_values.max_column + 1)]
        if any(v not in (None, "") for v in formula_vals):
            rows.append((r, cached_vals, formula_vals))
        if any(is_formula(v) for v in formula_vals):
            formula_rows.append(r)

    non_formula_data = [
        (r, cached, formulas)
        for r, cached, formulas in rows
        if any(v not in (None, "") and not is_formula(v) for v in formulas)
    ]

    col_stats = []
    for idx, header in enumerate(headers, start=1):
        vals = [cached[idx - 1] for _, cached, _ in rows]
        formula_vals = [formulas[idx - 1] for _, _, formulas in rows]
        filled = [v for v in vals if v not in (None, "")]
        formulas_count = sum(1 for v in formula_vals if is_formula(v))
        unique = Counter(cell_to_text(v) for v in filled)
        sample = []
        for item, _ in unique.most_common(6):
            if item and item not in sample:
                sample.append(item)
        numeric_vals = [v for v in filled if isinstance(v, (int, float)) and not isinstance(v, bool)]
        date_vals = [v for v in filled if isinstance(v, (datetime, date))]
        col_stats.append({
            "index": idx,
            "letter": openpyxl.utils.get_column_letter(idx),
            "header": header,
            "filled": len(filled),
            "formulas": formulas_count,
            "unique_count": len(unique),
            "sample": sample,
            "numeric": {
                "count": len(numeric_vals),
                "sum": round(sum(numeric_vals), 2) if numeric_vals else None,
                "min": min(numeric_vals) if numeric_vals else None,
                "max": max(numeric_vals) if numeric_vals else None,
            },
            "dates": {
                "count": len(date_vals),
                "min": min(date_vals).isoformat() if date_vals else None,
                "max": max(date_vals).isoformat() if date_vals else None,
            },
        })

    duplicate_candidates = []
    key_cols = []
    for i, h in enumerate(headers):
        hn = normalize_key(h)
        if any(token in hn for token in ["paciente", "cliente", "nome", "data", "dia", "horário", "horario"]):
            key_cols.append(i)
    if key_cols:
        seen = defaultdict(list)
        for r, cached, _ in non_formula_data:
            key = tuple(cell_to_text(cached[i]) for i in key_cols)
            if any(key):
                seen[key].append(r)
        duplicate_candidates = [{"key": k, "rows": v} for k, v in seen.items() if len(v) > 1][:20]

    formula_errors = []
    for r, _, formulas in rows:
        for idx, val in enumerate(formulas, start=1):
            if isinstance(val, str) and any(err in val for err in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]):
                formula_errors.append(f"{openpyxl.utils.get_column_letter(idx)}{r}: {val}")

    top_rows = []
    for r in range(1, min(ws_formulas.max_row, 15) + 1):
        row = compact_row([ws_formulas.cell(r, c).value for c in range(1, ws_formulas.max_column + 1)])
        if any(row):
            top_rows.append({"row": r, "values": row})

    return {
        "title": ws_formulas.title,
        "dimensions": {"rows": ws_formulas.max_row, "cols": ws_formulas.max_column},
        "header_row": header_row,
        "headers": headers,
        "rows_with_content": len(rows),
        "data_like_rows": len(non_formula_data),
        "rows_with_formulas": len(set(formula_rows)),
        "columns": col_stats,
        "duplicate_candidates": duplicate_candidates,
        "formula_errors": formula_errors,
        "top_rows": top_rows,
    }


def main():
    wb_values = openpyxl.load_workbook(INPUT, data_only=True)
    wb_formulas = openpyxl.load_workbook(INPUT, data_only=False)
    result = {"file": str(INPUT), "sheets": []}

    for name in wb_formulas.sheetnames:
        ws_v = wb_values[name]
        ws_f = wb_formulas[name]
        analysis = analyze_sheet(ws_v, ws_f)

        headers = analysis["headers"]
        header_row = analysis["header_row"]
        compact = {
            "title": name,
            "dimensions": analysis["dimensions"],
            "header_row": header_row,
            "headers_nonempty": [h for h in headers if not h.startswith("Coluna ")],
            "rows_with_content": analysis["rows_with_content"],
            "rows_with_formulas": analysis["rows_with_formulas"],
            "blank_trailing_columns": [
                c["letter"] for c in analysis["columns"]
                if c["filled"] == 0 and c["letter"] >= "J"
            ],
            "formula_errors": analysis["formula_errors"],
            "columns": [
                {
                    "letter": c["letter"],
                    "header": c["header"],
                    "filled": c["filled"],
                    "formulas": c["formulas"],
                    "unique_count": c["unique_count"],
                    "sample": c["sample"],
                    "numeric": c["numeric"],
                }
                for c in analysis["columns"]
                if c["filled"] or c["formulas"] or not c["header"].startswith("Coluna ")
            ],
        }

        status_counts = Counter()
        payment_status_counts = Counter()
        pay_mode_counts = Counter()
        qtd_payments_counts = Counter()
        total_missing_payment_status = 0
        rows_preview = []
        rows_flags = []
        formula_cells = []

        for r in range(header_row + 1, ws_f.max_row + 1):
            row_f = [ws_f.cell(r, c).value for c in range(1, 10)]
            row_v = [ws_v.cell(r, c).value for c in range(1, 10)]
            if not any(v not in (None, "") for v in row_f):
                continue
            values = [cell_to_text(v) for v in row_v]
            formulas = [cell_to_text(v) for v in row_f]
            name_value = values[0] or formulas[0]
            status = values[1] or formulas[1]
            pay_mode = values[2] or formulas[2]
            payment_status = values[4] or formulas[4]
            qtd_payments = values[5] or formulas[5]
            if status:
                status_counts[status] += 1
            if pay_mode:
                pay_mode_counts[pay_mode] += 1
            if payment_status:
                payment_status_counts[payment_status] += 1
            elif name_value:
                total_missing_payment_status += 1
            if qtd_payments:
                qtd_payments_counts[qtd_payments] += 1

            if len(rows_preview) < 40:
                rows_preview.append({
                    "row": r,
                    "A_I_values": values,
                    "A_I_formulas": formulas if any(is_formula(v) for v in row_f) else None,
                })

            if any(is_formula(v) for v in row_f):
                for idx, v in enumerate(row_f, start=1):
                    if is_formula(v):
                        formula_cells.append(f"{openpyxl.utils.get_column_letter(idx)}{r}")

            amount = row_v[3]
            if isinstance(amount, (int, float)) and amount >= 1000 and name_value:
                rows_flags.append({
                    "row": r,
                    "name": name_value,
                    "amount": amount,
                    "reason": "valor alto na coluna Valor da Sessão; pode ser total/resumo ou erro de lançamento",
                })
            if status == "Inativo" and payment_status and payment_status.upper().startswith("PEND"):
                rows_flags.append({
                    "row": r,
                    "name": name_value,
                    "reason": "paciente inativo aparece como pendente",
                })

        compact["status_counts"] = dict(status_counts)
        compact["payment_status_counts"] = dict(payment_status_counts)
        compact["payment_mode_counts"] = dict(pay_mode_counts)
        compact["qtd_payments_counts"] = dict(qtd_payments_counts)
        compact["missing_payment_status_rows"] = total_missing_payment_status
        compact["formula_cells_sample"] = formula_cells[:30]
        compact["flags"] = rows_flags[:40]
        compact["rows_preview"] = rows_preview
        result["sheets"].append(compact)

    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
